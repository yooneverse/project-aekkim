"""
애낌 합성 결제 데이터 생성 파이프라인 — 단일 유저용
-----------------------------------------------------
특징:
  - 쿠팡 제외 6개 OTT 중 3개 랜덤 선택
  - 선택된 각 OTT의 요금제 1개 랜덤 고정
  - 기간: 2025-01-01 ~ 2026-03-31 (15개월)
  - 구독 결제: 선택된 OTT 월 1회 고정 (결제일 N(base_day, std=2) 분포)
  - 비구독 결제: 월별 랜덤 건수 (15~90건) 생성
  - STEP 1: 데이터 생성 + 엑셀 저장 (카드 정보 불필요)
  - STEP 2: 실제 카드 결제 API 호출 (카드 정보 필요)
"""

import random
import datetime
import os
import numpy as np
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from pydantic import BaseModel
except ImportError:
    class BaseModel:
        def __init__(self, **kwargs): self.__dict__.update(kwargs)
        def model_dump(self): return self.__dict__

try:
    import requests
except ImportError:
    requests = None


# ──────────────────────────────────────────
# 섹션 1. 구독 서비스 정의 (쿠팡 제외)
# ──────────────────────────────────────────

SUBSCRIPTION_MERCHANTS = {
    "NETFLIX": [
        "NETFLIX_INICIS",
        "NETFLIX_INICIS-넷플릭스서비시스코리",
        "Netflix.com",
        "NETFLIX.COM",
    ],
    "TVING": [
        "주식회사 티빙",
    ],
    "WAVVE": [
        "웨이브",
        "Wavve",
        "콘텐츠웨이브㈜",
        "wavve자동결제",
    ],
    "DISNEY_PLUS": [
        "DISNEYPLUS",
        "디즈니플러스",
        "Disney+",
    ],
    "WATCHA": [
        "왓챠_스트리밍정기",
    ],
    # COUPANG_PLAY, NAVER_PLUS 제외
}

SERVICE_PLANS = {
    "NETFLIX": [
        {"plan": "광고형 스탠다드", "price": 7000},
        {"plan": "스탠다드",        "price": 13500},
        {"plan": "프리미엄",        "price": 17000},
    ],
    "TVING": [
        {"plan": "광고형 스탠다드", "price": 5500},
        {"plan": "스탠다드",        "price": 13500},
        {"plan": "프리미엄",        "price": 17000},
        {"plan": "베이직",        "price": 9500},
    ],
    "WAVVE": [
        {"plan": "광고형 스탠다드",   "price": 5500},
        {"plan": "베이직", "price": 7900},
        {"plan": "스탠다드", "price": 10900},
        {"plan": "프리미엄", "price": 13900},
    ],
    "DISNEY_PLUS": [
        {"plan": "스탠다드", "price": 9900},
        {"plan": "프리미엄", "price": 13900},
    ],
    "WATCHA": [
        {"plan": "베이직",   "price": 7900},
        {"plan": "프리미엄", "price": 12900},
    ],
}

# 서비스별 결제 기준일
SUBSCRIPTION_BASE_DAY = {
    "NETFLIX":     15,
    "TVING":       15,
    "WAVVE":       15,
    "DISNEY_PLUS": 15,
    "WATCHA":      15,
}

ALL_SERVICES = list(SUBSCRIPTION_MERCHANTS.keys())  # 5종 (쿠팡, 네이버플러스 제외)

# 생성 기간
START_DATE = datetime.date(2025, 1,  1)
END_DATE   = datetime.date(2026, 3, 31)


# ──────────────────────────────────────────
# 섹션 2. 비구독 가맹점명 풀
# ──────────────────────────────────────────

NON_SUB_POOL = {
    "카페·음료": {
        "merchants": [
            "텐퍼센트대연자이점", "스타벅스코리아", "하삼동커피부경대점",
            "하삼동커피녹산중앙", "(주)이스턴웰스본사", "텐퍼센트커피경성부",
            "커피베이부산오시리", "(주)이스턴웰스STX", "오오(55)커피",
            "매머드커피익스프레", "컴포즈커피부산롯데", "컴포즈커피범방점",
            "오빠다방", "코코노카",
        ],
        "count": 278,
    },
    "편의점·마트": {
        "merchants": [
            "씨유(CU)부산글로벌", "세븐일레븐부산대연", "보너스마트",
            "이마트24R부경대가", "탑플러스마트(부경대", "GS25부경삼광점",
            "이마트24대연유엔점", "씨유수정진성로점", "세븐일레븐장유덕정",
            "GS25부경대", "지에스(GS)25좌동코", "지에스25강서범방원",
        ],
        "count": 228,
    },
    "교통·이동": {
        "merchants": [
            "카카오T_바이크", "(주)이비카드택시법", "카카오페이(택시)",
            "SR", "카카오법인택시", "시외버스승차권",
            "이동의즐거움_택시_9", "코레일유통주식회사(", "고려주차장",
        ],
        "count": 127,
    },
    "쇼핑": {
        "merchants": [
            "네이버페이", "쿠팡(쿠페이)", "카카오선물하기_카카",
            "롯데몰동부산점매", "(주)이마트해운대점", "교보핫트랙스(주)",
            "(주)신세계백화점센", "베스토어", "카카오스타일",
        ],
        "count": 75,
    },
    "식비·외식": {
        "merchants": [
            "마츠도", "코코노카", "맥도날드부산송정DT",
            "푸드시스코리아(부", "데이앤데이미음점", "BKR버거킹부산대",
            "틈새라면", "마라패션(痲辣fashion", "할매분식",
            "맥도날드김해장유DT", "주식회사빡벳부경",
        ],
        "count": 71,
    },
    "학교": {
        "merchants": [
            "부경대학교소비자생", "부경대학교", "밀알케터링부경대점",
            "충북대학교", "에이스스터디카페",
        ],
        "count": 38,
    },
    "의료·건강": {
        "merchants": [
            "씨제이올리브네트웍", "율하신세계한의원", "(사)한국건강관리협",
            "해인온누리약국", "수미지이비인후과", "늘푸른약국",
            "명지명인한의원", "메디팜명성약국",
        ],
        "count": 34,
    },
    "베이커리·디저트": {
        "merchants": [
            "투썸플레이스장유율", "투썸플레이스더블유", "배스킨라빈스부산진",
            "(주)파리크라상경성", "배스킨라빈스", "뚜레주르장유젤미마",
        ],
        "count": 23,
    },
    "숙박": {
        "merchants": [
            "여우비호텔", "하버호텔", "(주)호텔신라인천공",
            "달토풀글램핑", "경포비치호텔", "호텔뮤리", "호텔스미스(SMITH)",
        ],
        "count": 10,
    },
    "세탁소": {
        "merchants": [
            "코인워시365대연자", "워시맨", "컴인워시부산본점", "크린토피아부산부경",
        ],
        "count": 10,
    },
    "여가·취미": {
        "merchants": [
            "좋아서하는", "세븐스타코인노래연", "씨씨와이코인노래연", "코인싱어동전노래연",
        ],
        "count": 9,
    },
    "꽃·선물": {
        "merchants": [
            "카르타플라워", "만우화원", "센츄리화원",
        ],
        "count": 6,
    },
    "기타": {
        "merchants": [
            "더-브릿지II(THEBRID", "오빠다방", "마츠도", "보너스마트",
            "KB국민카드_카카오페", "하삼동커피부경대점", "네이버페이",
            "탑플러스마트(부경대",
        ],
        "count": 234,
    },
}

_total_count       = sum(v["count"] for v in NON_SUB_POOL.values())
NON_SUB_CATEGORIES = list(NON_SUB_POOL.keys())
NON_SUB_WEIGHTS    = [v["count"] / _total_count for v in NON_SUB_POOL.values()]
assert abs(sum(NON_SUB_WEIGHTS) - 1.0) < 1e-9


def sample_non_sub_merchant() -> str:
    cat = random.choices(NON_SUB_CATEGORIES, weights=NON_SUB_WEIGHTS)[0]
    return random.choice(NON_SUB_POOL[cat]["merchants"])


# ──────────────────────────────────────────
# 섹션 3. 비구독 금액 분포
# ──────────────────────────────────────────

AMOUNT_BUCKETS = [
    (500,    10_000,  0.48),
    (10_000, 30_000,  0.40),
    (30_000, 150_000, 0.12),
]
_bucket_weights = [b[2] for b in AMOUNT_BUCKETS]


def sample_amount() -> int:
    lo, hi, _ = random.choices(AMOUNT_BUCKETS, weights=_bucket_weights)[0]
    mid = (lo + hi) / 2
    val = int(np.random.lognormal(np.log(mid), 0.4))
    return (max(lo, min(hi, val)) // 100) * 100


# ──────────────────────────────────────────
# 섹션 4. 유저 구독 프로필 선택
# ──────────────────────────────────────────

def build_subscription_profile() -> dict:
    """
    6개 서비스 중 3개 랜덤 선택 → 각 서비스별 가맹점 표기명·요금제 1개 고정.

    반환 예시:
    {
      "NETFLIX":     {"merchant": "Netflix.com",   "plan": "스탠다드", "price": 13500},
      "WAVVE":       {"merchant": "Wavve",          "plan": "베이직",   "price": 7900},
      "NAVER_PLUS":  {"merchant": "네이버플러스멤버십", "plan": "네이버플러스멤버십", "price": 4900},
    }
    """
    selected = random.sample(ALL_SERVICES, 3)
    profile  = {}
    for svc in selected:
        plan_obj = random.choice(SERVICE_PLANS[svc])
        profile[svc] = {
            "merchant": random.choice(SUBSCRIPTION_MERCHANTS[svc]),
            "plan":     plan_obj["plan"],
            "price":    plan_obj["price"],
        }
    return profile


# ──────────────────────────────────────────
# 섹션 5. 거래 내역 생성
# ──────────────────────────────────────────

def _iter_months(start: datetime.date, end: datetime.date):
    """start ~ end 기간의 (year, month) 순서대로 yield"""
    cur = start.replace(day=1)
    while cur <= end.replace(day=1):
        yield cur.year, cur.month
        cur += relativedelta(months=1)


def _sub_date(year: int, month: int, service_key: str) -> datetime.date:
    """서비스 기준일 기반 정규분포로 결제일 결정"""
    base_day = SUBSCRIPTION_BASE_DAY.get(service_key, 15)
    offset   = int(np.random.normal(0, 2))
    day      = max(1, min(28, base_day + offset))
    return datetime.date(year, month, day)


def generate_records(sub_profile: dict) -> list[dict]:
    """
    2025-01-01 ~ 2026-03-31 기간의 전체 거래 내역 생성.

    - 구독: 매월 선택된 3개 서비스 각 1건 (날짜 고정)
    - 비구독: 월별 랜덤 건수 (15~90건), 날짜 랜덤
    """
    records = []

    for year, month in _iter_months(START_DATE, END_DATE):
        # ── 구독 결제 (월 1회 고정) ──
        for svc_key, svc_info in sub_profile.items():
            date = _sub_date(year, month, svc_key)
            # 날짜가 해당 월 범위 초과 방어 (말일 처리)
            if date < START_DATE or date > END_DATE:
                continue
            records.append({
                "date":       date,
                "category":   "Subscription",
                "service":    svc_key,
                "merchant":   svc_info["merchant"],
                "plan":       svc_info["plan"],
                "amount_krw": svc_info["price"],
            })

        # ── 비구독 결제 (월별 랜덤 건수) ──
        import calendar
        last_day = calendar.monthrange(year, month)[1]
        month_start = datetime.date(year, month, 1)
        month_end   = min(datetime.date(year, month, last_day), END_DATE)
        n_nonsub    = random.randint(15, 90)

        for _ in range(n_nonsub):
            delta = random.randint(0, (month_end - month_start).days)
            date  = month_start + datetime.timedelta(days=delta)
            records.append({
                "date":       date,
                "category":   "NonSub",
                "service":    None,
                "merchant":   sample_non_sub_merchant(),
                "plan":       None,
                "amount_krw": sample_amount(),
            })

    # 날짜 오름차순 정렬
    records.sort(key=lambda r: r["date"])
    return records


# ──────────────────────────────────────────
# 섹션 6. 엑셀 출력
# ──────────────────────────────────────────

def export_to_excel(records: list[dict], sub_profile: dict, output_path: str) -> None:
    wb = Workbook()

    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill_b = PatternFill("solid", fgColor="1F4E79")
    header_fill_l = PatternFill("solid", fgColor="2E75B6")
    sub_fill      = PatternFill("solid", fgColor="E2EFDA")
    alt_fill      = PatternFill("solid", fgColor="F2F2F2")
    center        = Alignment(horizontal="center", vertical="center")
    left          = Alignment(horizontal="left",   vertical="center")
    right         = Alignment(horizontal="right",  vertical="center")
    thin          = Side(style="thin", color="CCCCCC")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font     = Font(name="Arial", size=10)

    def style_header(cell, fill=header_fill_b):
        cell.font = header_font; cell.fill = fill
        cell.alignment = center; cell.border = border

    def style_cell(cell, align=left, fill=None):
        cell.font = body_font; cell.alignment = align; cell.border = border
        if fill: cell.fill = fill

    # ════════════════════════════════
    # 시트 1: 전체 거래 내역
    # ════════════════════════════════
    ws1 = wb.active
    ws1.title = "거래 내역"

    headers    = ["No", "거래일", "카테고리", "서비스", "요금제",
                  "가맹점명(merchantName)", "금액(원)"]
    col_widths = [6, 14, 12, 16, 20, 34, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=h)
        style_header(cell)
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[1].height = 20

    for i, rec in enumerate(records, 1):
        is_sub   = rec["category"] == "Subscription"
        row_fill = sub_fill if is_sub else (alt_fill if i % 2 == 0 else None)
        vals     = [i,
                    rec["date"].strftime("%Y-%m-%d"),
                    rec["category"],
                    rec["service"] or "",
                    rec["plan"]    or "",
                    rec["merchant"],
                    rec["amount_krw"]]
        aligns   = [center, center, center, center, left, left, right]

        for col, (val, aln) in enumerate(zip(vals, aligns), 1):
            cell = ws1.cell(row=i + 1, column=col, value=val)
            style_cell(cell, align=aln, fill=row_fill)
            if col == 7:
                cell.number_format = '#,##0'

    ws1.auto_filter.ref = f"A1:G{len(records)+1}"
    ws1.freeze_panes    = "A2"

    # ════════════════════════════════
    # 시트 2: 구독 프로필 요약
    # ════════════════════════════════
    ws2 = wb.create_sheet("구독 프로필")

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 34
    ws2.column_dimensions["E"].width = 14

    prof_headers = ["서비스", "요금제", "월 결제금액(원)", "가맹점명(merchantName)", "결제 기준일"]
    for col, h in enumerate(prof_headers, 1):
        style_header(ws2.cell(row=1, column=col, value=h), fill=header_fill_l)

    for i, (svc_key, info) in enumerate(sub_profile.items(), 2):
        vals = [svc_key, info["plan"], info["price"],
                info["merchant"], f"{SUBSCRIPTION_BASE_DAY[svc_key]}일 전후"]
        aligns = [center, left, right, left, center]
        for col, (val, aln) in enumerate(zip(vals, aligns), 1):
            cell = ws2.cell(row=i, column=col, value=val)
            style_cell(cell, align=aln, fill=alt_fill if i % 2 == 0 else None)
            if col == 3:
                cell.number_format = '#,##0'

    # ════════════════════════════════
    # 시트 3: 월별 통계
    # ════════════════════════════════
    ws3 = wb.create_sheet("월별 통계")
    df  = pd.DataFrame(records)
    df["month"] = df["date"].apply(lambda d: d.strftime("%Y-%m"))

    monthly = (df.groupby(["month", "category"])
                 .agg(건수=("amount_krw", "count"), 합계=("amount_krw", "sum"))
                 .reset_index())

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 16

    stat_headers = ["월", "카테고리", "건수", "합계(원)"]
    for col, h in enumerate(stat_headers, 1):
        style_header(ws3.cell(row=1, column=col, value=h), fill=header_fill_l)

    for i, row in enumerate(monthly.itertuples(), 2):
        vals   = [row.month, row.category, row.건수, row.합계]
        aligns = [center, center, center, right]
        row_fill = sub_fill if row.category == "Subscription" else (alt_fill if i % 2 == 0 else None)
        for col, (val, aln) in enumerate(zip(vals, aligns), 1):
            cell = ws3.cell(row=i, column=col, value=val)
            style_cell(cell, align=aln, fill=row_fill)
            if col == 4:
                cell.number_format = '#,##0'

    wb.save(output_path)
    print(f"    엑셀 저장 완료: {output_path}")


# ──────────────────────────────────────────
# 섹션 7. SSAFY 카드 결제 API 형식
# ──────────────────────────────────────────

class SSAFYHeader(BaseModel):
    apiName:                        str
    transmissionDate:               str
    transmissionTime:               str
    institutionCode:                str = "00100"
    fintechAppNo:                   str = "001"
    apiServiceCode:                 str
    institutionTransactionUniqueNo: str
    apiKey:                         str
    userKey:                        str


class CardTransactionRequest(BaseModel):
    Header:         SSAFYHeader
    cardNo:         str
    cvc:            str
    merchantId:     str
    paymentBalance: str


_seq_counter = 0

def _unique_no() -> str:
    global _seq_counter
    _seq_counter += 1
    return datetime.datetime.now().strftime("%Y%m%d%H%M%S") + str(_seq_counter).zfill(6)


def to_card_request(record: dict, card_no: str, cvc: str,
                    merchant_id: str, api_key: str, user_key: str) -> CardTransactionRequest:
    txn_date = record["date"]
    date_str = txn_date.strftime("%Y%m%d")
    time_str = (f"{random.randint(0,23):02d}"
                f"{random.randint(0,59):02d}"
                f"{random.randint(0,59):02d}")
    header = SSAFYHeader(
        apiName="createCreditCardTransaction",
        transmissionDate=date_str,
        transmissionTime=time_str,
        apiServiceCode="createCreditCardTransaction",
        institutionTransactionUniqueNo=_unique_no(),
        apiKey=api_key,
        userKey=user_key,
    )
    return CardTransactionRequest(
        Header=header,
        cardNo=card_no,
        cvc=cvc,
        merchantId=merchant_id,
        paymentBalance=str(record["amount_krw"]),
    )


# ──────────────────────────────────────────
# 섹션 8. API 호출
# ──────────────────────────────────────────

API_URL = os.getenv(
    "FINANCE_CARD_API_URL",
    "https://finance-api.example.com/v1/credit-card/transactions",
)

def call_card_transaction(req: CardTransactionRequest) -> dict:
    resp = requests.post(API_URL, json=req.model_dump(), timeout=10)
    resp.raise_for_status()
    return resp.json()


# ──────────────────────────────────────────
# 섹션 9. 메인 파이프라인
# ──────────────────────────────────────────

def run_pipeline(
    generate_only: bool = True,
    excel_path:    str  = "single_user_transactions.xlsx",
    card_no:       str  = "",
    cvc:           str  = "",
    merchant_id:   str  = "",
    api_key:       str  = "",
    user_key:      str  = "",
):
    """
    generate_only=True  → 데이터 생성 + 엑셀 저장만
    generate_only=False → 엑셀 저장 후 실제 카드 결제 API 호출
    """
    print("[1] 구독 프로필 선택 ...")
    sub_profile = build_subscription_profile()
    for svc, info in sub_profile.items():
        print(f"    {svc}: {info['plan']} ({info['price']:,}원) / {info['merchant']}")

    print(f"[2] 거래 내역 생성 ({START_DATE} ~ {END_DATE}) ...")
    records = generate_records(sub_profile)
    sub_cnt = sum(1 for r in records if r["category"] == "Subscription")
    print(f"    총 {len(records)}건 (구독 {sub_cnt}건 / 비구독 {len(records)-sub_cnt}건)")

    print(f"[3] 엑셀 저장 ({excel_path}) ...")
    export_to_excel(records, sub_profile, excel_path)

    if generate_only:
        print("\n[생성 완료] generate_only=True 모드")
        print("  → 엑셀에서 데이터 확인 후 아래 항목 채워서 재실행하세요:")
        print("      card_no, cvc, merchant_id, api_key, user_key")
        print("      generate_only=False")
        return records

    # ── API 호출 모드 ──
    if not all([card_no, cvc, merchant_id, api_key, user_key]):
        raise ValueError("generate_only=False 시 card_no, cvc, merchant_id, api_key, user_key 모두 필요")

    print(f"[4] 카드 결제 API 호출 ({len(records)}건) ...")
    ok, fail = 0, 0
    for record in records:
        req = to_card_request(record, card_no, cvc, merchant_id, api_key, user_key)
        try:
            call_card_transaction(req)
            ok += 1
        except Exception as e:
            fail += 1
            if fail <= 5:
                print(f"    [FAIL] {record['merchant']}: {e}")

    print(f"[완료] 성공 {ok}건 / 실패 {fail}건")
    return records


# ──────────────────────────────────────────
# 실행
# ──────────────────────────────────────────

if __name__ == "__main__":

    # ── STEP 1: 데이터 생성 + 엑셀 저장 ──
    run_pipeline(
        generate_only=True,
        excel_path="single_user_transactions.xlsx",
    )

    # ── STEP 2: 백엔드에서 카드 정보 받은 후 아래 주석 해제하고 재실행 ──
    # run_pipeline(
    #     generate_only=False,
    #     excel_path="single_user_transactions.xlsx",
    #     card_no="",        # 백엔드에서 받은 카드번호 (16자리)
    #     cvc="",            # CVC (3자리)
    #     merchant_id="",    # SSAFY 가맹점 ID
    #     api_key="",        # SSAFY API 키
    #     user_key="",       # SSAFY 유저 키
    # )

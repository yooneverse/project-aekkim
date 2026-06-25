"""
애낌 합성 결제 데이터 생성 파이프라인 v4.0
--------------------------------------------
변경 이력 (v3.0 → v4.0):
  [1] SUBSCRIPTION_MERCHANTS 키 영문화 (NETFLIX, TVING, ...)
  [2] SERVICE_PLANS 추가 — 서비스별 요금제 + 고정 금액 정의
  [3] 유저 프로필 도입 — 유저 생성 시 서비스별 가맹점 표기명·요금제 1개씩 고정
  [4] ① 버그 수정: 구독 수 카운팅을 서비스 단위로 교정 (표기명 단위 → 서비스 단위)
  [5] 케이스 확장: A/B/C/D(고정) + E(요금제 변경) + F(구독 추가) 총 6종
  [6] 티빙+디즈니번들 제외 (추후 추가 예정)
  [7] n_users 기본값 60으로 상향 (케이스 6종 × 10명 균등 보장)

실행 모드:
  - generate_only=True  : 카드 정보 없이 데이터 생성 + 엑셀 저장
  - generate_only=False : 카드 정보 입력 후 실제 API 호출
"""

import random
import datetime
import json
import os
import numpy as np
import pandas as pd
from collections import Counter
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from pydantic import BaseModel
except ImportError:
    class BaseModel:
        def __init__(self, **kwargs): self.__dict__.update(kwargs)
        def model_dump(self): return self.__dict__

try:
    import requests
except ImportError:
    requests = None


# ──────────────────────────────────────────
# 섹션 1. 구독 서비스 정의
# ──────────────────────────────────────────

# 서비스 키(영문) → 실제 카드 결제창 표기명 리스트
# 유저 프로필 생성 시 표기명 1개를 랜덤으로 고정해 사용
SUBSCRIPTION_MERCHANTS = {
    "NETFLIX": [
        "NETFLIX_INICIS",                     # 실제 조사
        "NETFLIX_INICIS-넷플릭스서비시스코리",   # 실제 조사
        "Netflix.com",                        # 실제 조사
        "NETFLIX.COM",                        # 실제 조사
    ],
    "TVING": [
        "주식회사 티빙",                         # 실제 조사
    ],
    "WAVVE": [
        "웨이브",                               # 웹 검색 결과
        "Wavve",                              # 웹 검색 결과
        "콘텐츠웨이브㈜",                         # 웹 검색 결과
        "wavve자동결제",                         # 웹 검색 결과
    ],
    "DISNEY_PLUS": [
        "DISNEYPLUS",                         # 웹 검색 결과
        "디즈니플러스",                          # 웹 검색 결과
        "Disney+",                            # 웹 검색 결과
    ],
    "WATCHA": [
        "왓챠_스트리밍정기",                       # 실제 조사
    ],
    "COUPANG_PLAY": [
        "쿠팡와우월회비",                         # 실제 조사
        "쿠팡(와우멤버십)",                       # 실제 조사
    ],
    "NAVER_PLUS": [
        "네이버플러스멤버십",                       # 실제 조사
    ],
    # 티빙+디즈니번들: 추후 추가 예정
}

# 서비스 키(영문) → 요금제 목록 [{plan, price}, ...]
# 구독 결제 금액은 항상 이 테이블에서만 참조 (랜덤 샘플링 금지)
SERVICE_PLANS = {
    "NETFLIX": [
        {"plan": "광고형 스탠다드", "price": 4900},
        {"plan": "스탠다드",        "price": 13500},
        {"plan": "프리미엄",        "price": 17000},
    ],
    "TVING": [
        {"plan": "광고형 스탠다드", "price": 4900},
        {"plan": "스탠다드",        "price": 10900},
        {"plan": "프리미엄",        "price": 13900},
    ],
    "WAVVE": [
        {"plan": "베이직",   "price": 7900},
        {"plan": "스탠다드", "price": 10900},
        {"plan": "프리미엄", "price": 13900},
    ],
    "DISNEY_PLUS": [
        {"plan": "스탠다드", "price": 9900},
        {"plan": "프리미엄", "price": 13900},
    ],
    "WATCHA": [
        {"plan": "베이직",   "price": 7900},
        {"plan": "프리미엄", "price": 12900},
    ],
    "COUPANG_PLAY": [
        {"plan": "와우멤버십", "price": 7890},
    ],
    "NAVER_PLUS": [
        {"plan": "네이버플러스멤버십", "price": 4900},
    ],
}

# 서비스별 결제 기준일 (정규분포 N(base_day, std=2.0) 로 날짜 샘플링)
SUBSCRIPTION_BASE_DAY = {
    "NETFLIX":      15,
    "TVING":        15,
    "WAVVE":        15,
    "DISNEY_PLUS":  15,
    "WATCHA":       15,
    "COUPANG_PLAY": 20,
    "NAVER_PLUS":   15,
}

# 전체 서비스 목록 (케이스 샘플링용)
ALL_SERVICES = list(SUBSCRIPTION_MERCHANTS.keys())


# ──────────────────────────────────────────
# 섹션 2. 비구독 가맹점명 풀 (실측 건수 기반 정규화 가중치)
# ──────────────────────────────────────────

NON_SUB_POOL = {
    "카페·음료": {
        "merchants": [
            "텐퍼센트대연자이점", "스타벅스코리아", "하삼동커피부경대점",
            "하삼동커피녹산중앙", "(주)이스턴웰스본사", "텐퍼센트커피경성부",
            "커피베이부산오시리", "(주)이스턴웰스STX", "오오(55)커피",
            "매머드커피익스프레", "컴포즈커피부산롯데", "컴포즈커피범방점",
            "오빠다방", "코코노카",
        ],
        "count": 278,
    },
    "편의점·마트": {
        "merchants": [
            "씨유(CU)부산글로벌", "세븐일레븐부산대연", "보너스마트",
            "이마트24R부경대가", "탑플러스마트(부경대", "GS25부경삼광점",
            "이마트24대연유엔점", "씨유수정진성로점", "세븐일레븐장유덕정",
            "GS25부경대", "지에스(GS)25좌동코", "지에스25강서범방원",
        ],
        "count": 228,
    },
    "교통·이동": {
        "merchants": [
            "카카오T_바이크", "(주)이비카드택시법", "카카오페이(택시)",
            "SR", "카카오법인택시", "시외버스승차권",
            "이동의즐거움_택시_9", "코레일유통주식회사(", "고려주차장",
        ],
        "count": 127,
    },
    "쇼핑": {
        "merchants": [
            "네이버페이", "쿠팡(쿠페이)", "카카오선물하기_카카",
            "롯데몰동부산점매", "(주)이마트해운대점", "교보핫트랙스(주)",
            "(주)신세계백화점센", "베스토어", "카카오스타일",
        ],
        "count": 75,
    },
    "식비·외식": {
        "merchants": [
            "마츠도", "코코노카", "맥도날드부산송정DT",
            "푸드시스코리아(부", "데이앤데이미음점", "BKR버거킹부산대",
            "틈새라면", "마라패션(痲辣fashion", "할매분식",
            "맥도날드김해장유DT", "주식회사빡벳부경",
        ],
        "count": 71,
    },
    "학교": {
        "merchants": [
            "부경대학교소비자생", "부경대학교", "밀알케터링부경대점",
            "충북대학교", "에이스스터디카페",
        ],
        "count": 38,
    },
    "의료·건강": {
        "merchants": [
            "씨제이올리브네트웍", "율하신세계한의원", "(사)한국건강관리협",
            "해인온누리약국", "수미지이비인후과", "늘푸른약국",
            "명지명인한의원", "메디팜명성약국",
        ],
        "count": 34,
    },
    "베이커리·디저트": {
        "merchants": [
            "투썸플레이스장유율", "투썸플레이스더블유", "배스킨라빈스부산진",
            "(주)파리크라상경성", "배스킨라빈스", "뚜레주르장유젤미마",
        ],
        "count": 23,
    },
    "숙박": {
        "merchants": [
            "여우비호텔", "하버호텔", "(주)호텔신라인천공",
            "달토풀글램핑", "경포비치호텔", "호텔뮤리", "호텔스미스(SMITH)",
        ],
        "count": 10,
    },
    "세탁소": {
        "merchants": [
            "코인워시365대연자", "워시맨", "컴인워시부산본점", "크린토피아부산부경",
        ],
        "count": 10,
    },
    "여가·취미": {
        "merchants": [
            "좋아서하는", "세븐스타코인노래연", "씨씨와이코인노래연", "코인싱어동전노래연",
        ],
        "count": 9,
    },
    "꽃·선물": {
        "merchants": [
            "카르타플라워", "만우화원", "센츄리화원",
        ],
        "count": 6,
    },
    "기타": {
        "merchants": [
            "더-브릿지II(THEBRID", "오빠다방", "마츠도", "보너스마트",
            "KB국민카드_카카오페", "하삼동커피부경대점", "네이버페이",
            "탑플러스마트(부경대",
        ],
        "count": 234,
    },
}

_total_count       = sum(v["count"] for v in NON_SUB_POOL.values())
NON_SUB_CATEGORIES = list(NON_SUB_POOL.keys())
NON_SUB_WEIGHTS    = [v["count"] / _total_count for v in NON_SUB_POOL.values()]
assert abs(sum(NON_SUB_WEIGHTS) - 1.0) < 1e-9, "비구독 가중치 합 오류"


def sample_non_sub_merchant() -> str:
    cat = random.choices(NON_SUB_CATEGORIES, weights=NON_SUB_WEIGHTS)[0]
    return random.choice(NON_SUB_POOL[cat]["merchants"])


# ──────────────────────────────────────────
# 섹션 3. 금액 분포 (비구독 전용)
# ──────────────────────────────────────────

AMOUNT_BUCKETS = [
    (500,    10_000,  0.48),
    (10_000, 30_000,  0.40),
    (30_000, 150_000, 0.12),
]
_bucket_weights = [b[2] for b in AMOUNT_BUCKETS]
assert abs(sum(_bucket_weights) - 1.0) < 1e-9, "금액 버킷 가중치 합 오류"


def sample_amount() -> int:
    """비구독 거래 금액 샘플링 (로그정규분포)"""
    lo, hi, _ = random.choices(AMOUNT_BUCKETS, weights=_bucket_weights)[0]
    mid = (lo + hi) / 2
    val = int(np.random.lognormal(np.log(mid), 0.4))
    val = max(lo, min(hi, val))
    return (val // 100) * 100


def generate_base_transactions(n_users: int) -> pd.DataFrame:
    rows = []
    base_date = datetime.date(2024, 1, 1)
    for user_id in range(1, n_users + 1):
        n_txn = random.randint(15, 90)
        for _ in range(n_txn):
            delta = random.randint(0, 364)
            date  = base_date + datetime.timedelta(days=delta)
            cat   = random.choices(
                ["Subscription", "NonSub"], weights=[0.15, 0.85]
            )[0]
            rows.append({"user_id": user_id, "date": date, "category": cat})
    return pd.DataFrame(rows)


# ──────────────────────────────────────────
# 섹션 4. 유저 케이스 정의 및 프로필 생성
# ──────────────────────────────────────────

CASE_LABELS = {
    "A": "구독 1개, 요금제 고정",
    "B": "구독 2개, 요금제 고정",
    "C": "구독 3개, 요금제 고정",
    "D": "구독 4~6개, 요금제 고정",
    "E": "구독 1개, 중간에 요금제 변경",
    "F": "구독 N개 → N+1개 (신규 구독 추가)",
}

ALL_CASES = ["A", "B", "C", "D", "E", "F"]


def _pick_service_profile(service_key: str) -> dict:
    """서비스 키 → {merchant, plan, price} 고정 프로필 생성"""
    merchant = random.choice(SUBSCRIPTION_MERCHANTS[service_key])
    plan_obj = random.choice(SERVICE_PLANS[service_key])
    return {
        "merchant": merchant,
        "plan":     plan_obj["plan"],
        "price":    plan_obj["price"],
    }


def build_user_profile(case: str) -> dict:
    """
    유저 1명의 구독 프로필을 생성한다.

    서비스 선택 → 표기명 1개 고정 → 요금제 1개 고정 순서로 처리.
    케이스 E/F는 시간축 변화 관련 추가 필드를 포함한다.

    반환 예시 (케이스 E):
    {
      "case": "E",
      "services": {
        "NETFLIX": {"merchant": "Netflix.com", "plan": "스탠다드", "price": 13500}
      },
      "change_month": 6,
      "change_service": "NETFLIX",
      "plan_before": {"plan": "스탠다드", "price": 13500},
      "plan_after":  {"plan": "프리미엄", "price": 17000},
    }
    """
    profile: dict = {"case": case, "services": {}}

    if case == "A":
        svc = random.sample(ALL_SERVICES, 1)
        for s in svc:
            profile["services"][s] = _pick_service_profile(s)

    elif case == "B":
        svc = random.sample(ALL_SERVICES, 2)
        for s in svc:
            profile["services"][s] = _pick_service_profile(s)

    elif case == "C":
        svc = random.sample(ALL_SERVICES, 3)
        for s in svc:
            profile["services"][s] = _pick_service_profile(s)

    elif case == "D":
        n   = random.randint(4, min(6, len(ALL_SERVICES)))
        svc = random.sample(ALL_SERVICES, n)
        for s in svc:
            profile["services"][s] = _pick_service_profile(s)

    elif case == "E":
        # 구독 1개, 중간에 요금제 변경
        svc_key  = random.choice(ALL_SERVICES)
        plans    = SERVICE_PLANS[svc_key]
        merchant = random.choice(SUBSCRIPTION_MERCHANTS[svc_key])

        if len(plans) >= 2:
            before, after = random.sample(plans, 2)
        else:
            # 요금제 1개뿐인 서비스 엣지케이스 방어
            before = after = plans[0]

        profile["services"][svc_key] = {
            "merchant": merchant,
            "plan":     before["plan"],
            "price":    before["price"],
        }
        profile["change_month"]   = random.randint(3, 9)
        profile["change_service"] = svc_key
        profile["plan_before"]    = {"plan": before["plan"], "price": before["price"]}
        profile["plan_after"]     = {"plan": after["plan"],  "price": after["price"]}

    elif case == "F":
        # 기준 N개 구독 → 특정 월부터 N+1개
        n_base    = random.randint(1, min(3, len(ALL_SERVICES) - 1))
        base_svcs = random.sample(ALL_SERVICES, n_base)
        remaining = [s for s in ALL_SERVICES if s not in base_svcs]
        added_svc = random.choice(remaining)

        for s in base_svcs:
            profile["services"][s] = _pick_service_profile(s)

        profile["add_month"]     = random.randint(3, 9)
        profile["added_service"] = added_svc
        profile["added_profile"] = _pick_service_profile(added_svc)

    return profile


def build_all_user_profiles(n_users: int) -> dict:
    """n_users명의 유저 프로필 딕셔너리 반환 {uid: profile}"""
    # 6종 케이스 순환 후 셔플 → 케이스당 n_users // 6 명 보장
    case_cycle = [ALL_CASES[i % len(ALL_CASES)] for i in range(n_users)]
    random.shuffle(case_cycle)
    return {uid: build_user_profile(case_cycle[uid - 1])
            for uid in range(1, n_users + 1)}


# ──────────────────────────────────────────
# 섹션 5. 후처리 — 구독 결제 생성
# ──────────────────────────────────────────

def _make_sub_day(service_key: str) -> int:
    """서비스 기준일 기반 정규분포로 결제일 샘플링"""
    base_day = SUBSCRIPTION_BASE_DAY.get(service_key, 15)
    offset   = int(np.random.normal(0, 2))
    return max(1, min(28, base_day + offset))


def _make_nonsub_record(uid: int, date: datetime.date, profile: dict) -> dict:
    return {
        "user_id":    uid,
        "user_case":  profile["case"],
        "case_label": CASE_LABELS[profile["case"]],
        "date":       date,
        "category":   "NonSub",
        "service":    None,
        "merchant":   sample_non_sub_merchant(),
        "plan":       None,
        "amount_krw": sample_amount(),
    }


def postprocess(base_df: pd.DataFrame, user_profiles: dict) -> list[dict]:
    n_users  = len(user_profiles)
    records  = []
    seen_sub: dict = {}  # (uid, service_key, YYYY-MM) → 월 중복 방지

    for _, row in base_df.iterrows():
        uid      = max(1, min(int(row.get("user_id", 1)), n_users))
        date     = row["date"] if isinstance(row["date"], datetime.date) \
                   else datetime.date.today()
        category = row.get("category", "NonSub")
        profile  = user_profiles[uid]
        case     = profile["case"]

        if category == "Subscription":
            # ── 현재 달에 활성화된 서비스 목록 결정 ──
            month          = date.month
            active_services = dict(profile["services"])  # 기본 구독 복사

            if case == "E":
                # 변경 시점 이후: 해당 서비스 요금제를 after로 교체
                svc_key = profile["change_service"]
                if month >= profile["change_month"]:
                    active_services[svc_key] = {
                        "merchant": profile["services"][svc_key]["merchant"],
                        **profile["plan_after"],
                    }

            elif case == "F":
                # 추가 시점 이후: added_service 포함
                if month >= profile["add_month"]:
                    active_services[profile["added_service"]] = profile["added_profile"]

            if not active_services:
                records.append(_make_nonsub_record(uid, date, profile))
                continue

            # ── 서비스 1개 선택 (서비스 단위 중복 방지) ──
            service_key = random.choice(list(active_services.keys()))
            month_key   = (uid, service_key, date.strftime("%Y-%m"))

            if month_key in seen_sub:
                # 이 달에 이미 해당 서비스 결제 → 비구독 대체
                records.append(_make_nonsub_record(uid, date, profile))
                continue

            seen_sub[month_key] = True
            svc_profile = active_services[service_key]

            # ── 결제일 집중 로직 ──
            new_day = _make_sub_day(service_key)
            date    = date.replace(day=new_day)

            records.append({
                "user_id":    uid,
                "user_case":  case,
                "case_label": CASE_LABELS[case],
                "date":       date,
                "category":   "Subscription",
                "service":    service_key,
                "merchant":   svc_profile["merchant"],
                "plan":       svc_profile["plan"],
                "amount_krw": svc_profile["price"],
            })

        else:
            records.append(_make_nonsub_record(uid, date, profile))

    return records


# ──────────────────────────────────────────
# 섹션 6. 엑셀 출력
# ──────────────────────────────────────────

def export_to_excel(records: list[dict], output_path: str) -> None:
    """
    시트 1: 전체 거래 내역
    시트 2: 유저별 요약
    시트 3: 통계 요약
    """
    wb = Workbook()

    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill_b = PatternFill("solid", fgColor="1F4E79")
    header_fill_l = PatternFill("solid", fgColor="2E75B6")
    sub_fill      = PatternFill("solid", fgColor="E2EFDA")   # 연초록 (구독)
    case_e_fill   = PatternFill("solid", fgColor="FCE4D6")   # 연주황 (케이스 E)
    case_f_fill   = PatternFill("solid", fgColor="EAD1DC")   # 연보라 (케이스 F)
    alt_fill      = PatternFill("solid", fgColor="F2F2F2")
    center        = Alignment(horizontal="center", vertical="center")
    left          = Alignment(horizontal="left",   vertical="center")
    right         = Alignment(horizontal="right",  vertical="center")
    thin          = Side(style="thin", color="CCCCCC")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font     = Font(name="Arial", size=10)

    def style_header(cell, fill=header_fill_b):
        cell.font = header_font; cell.fill = fill
        cell.alignment = center; cell.border = border

    def style_cell(cell, align=left, fill=None):
        cell.font = body_font; cell.alignment = align; cell.border = border
        if fill:
            cell.fill = fill

    # ════════════════════════════════
    # 시트 1: 전체 거래 내역
    # ════════════════════════════════
    ws1 = wb.active
    ws1.title = "거래 내역"

    headers    = ["No", "유저ID", "케이스", "케이스 설명",
                  "거래일", "카테고리", "서비스", "요금제",
                  "가맹점명(merchantName)", "금액(원)"]
    col_widths = [6, 8, 8, 30, 14, 12, 16, 18, 32, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=h)
        style_header(cell)
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[1].height = 20

    for i, rec in enumerate(records, 1):
        is_sub   = rec["category"] == "Subscription"
        c        = rec["user_case"]
        if is_sub:
            row_fill = sub_fill
        elif c == "E":
            row_fill = case_e_fill
        elif c == "F":
            row_fill = case_f_fill
        elif i % 2 == 0:
            row_fill = alt_fill
        else:
            row_fill = None

        vals   = [i, rec["user_id"], c, rec["case_label"],
                  rec["date"].strftime("%Y-%m-%d"), rec["category"],
                  rec["service"] or "", rec["plan"] or "",
                  rec["merchant"], rec["amount_krw"]]
        aligns = [center, center, center, left,
                  center, center, center, left,
                  left, right]

        for col, (val, aln) in enumerate(zip(vals, aligns), 1):
            cell = ws1.cell(row=i + 1, column=col, value=val)
            style_cell(cell, align=aln, fill=row_fill)
            if col == 10:
                cell.number_format = '#,##0'

    ws1.auto_filter.ref = f"A1:J{len(records)+1}"
    ws1.freeze_panes    = "A2"

    # ════════════════════════════════
    # 시트 2: 유저별 요약
    # ════════════════════════════════
    ws2 = wb.create_sheet("유저별 요약")
    df  = pd.DataFrame(records)

    summary_rows = []
    for uid in sorted(df["user_id"].unique()):
        udf       = df[df["user_id"] == uid]
        case      = udf["user_case"].iloc[0]
        label     = udf["case_label"].iloc[0]
        total     = len(udf)
        sub_cnt   = (udf["category"] == "Subscription").sum()
        sub_svcs  = ", ".join(sorted(
            udf[udf["category"] == "Subscription"]["service"].dropna().unique()
        ))
        sub_names = ", ".join(sorted(
            udf[udf["category"] == "Subscription"]["merchant"].unique()
        ))
        total_amt = udf["amount_krw"].sum()
        summary_rows.append([uid, case, label, total, sub_cnt, sub_svcs, sub_names, total_amt])

    s2_headers = ["유저ID", "케이스", "케이스 설명",
                  "총 거래 건수", "구독 거래 건수",
                  "구독 서비스(영문키)", "구독 가맹점명", "총 지출액(원)"]
    s2_widths  = [8, 8, 28, 14, 14, 30, 50, 16]

    for col, (h, w) in enumerate(zip(s2_headers, s2_widths), 1):
        cell = ws2.cell(row=1, column=col, value=h)
        style_header(cell, fill=header_fill_l)
        ws2.column_dimensions[get_column_letter(col)].width = w

    for i, row_data in enumerate(summary_rows, 2):
        c        = row_data[1]
        row_fill = (case_e_fill if c == "E" else
                    case_f_fill if c == "F" else
                    alt_fill    if i % 2 == 0 else None)
        aligns_s = [center, center, left, center, center, left, left, right]
        for col, (val, aln) in enumerate(zip(row_data, aligns_s), 1):
            cell = ws2.cell(row=i, column=col, value=val)
            style_cell(cell, align=aln, fill=row_fill)
            if col == 8:
                cell.number_format = '#,##0'

    ws2.auto_filter.ref = f"A1:H{len(summary_rows)+1}"
    ws2.freeze_panes    = "A2"

    # ════════════════════════════════
    # 시트 3: 통계 요약
    # ════════════════════════════════
    ws3 = wb.create_sheet("통계 요약")
    amounts = df["amount_krw"]

    stats = [
        ["항목", "값"],
        ["총 거래 건수", len(records)],
        ["총 유저 수", df["user_id"].nunique()],
        ["구독 거래 건수", (df["category"] == "Subscription").sum()],
        ["비구독 거래 건수", (df["category"] == "NonSub").sum()],
        ["", ""],
        ["금액 통계", ""],
        ["금액 중앙값(원)", int(amounts.median())],
        ["금액 평균(원)", int(amounts.mean())],
        ["금액 최소(원)", int(amounts.min())],
        ["금액 최대(원)", int(amounts.max())],
        ["소액(≤10,000원) 비중", f"{(amounts <= 10000).mean() * 100:.1f}%"],
        ["", ""],
        ["케이스 분포 (거래 건수 기준)", ""],
        ["케이스 A (구독 1개, 고정)", (df["user_case"] == "A").sum()],
        ["케이스 B (구독 2개, 고정)", (df["user_case"] == "B").sum()],
        ["케이스 C (구독 3개, 고정)", (df["user_case"] == "C").sum()],
        ["케이스 D (구독 4~6개, 고정)", (df["user_case"] == "D").sum()],
        ["케이스 E (요금제 변경)", (df["user_case"] == "E").sum()],
        ["케이스 F (구독 추가)", (df["user_case"] == "F").sum()],
        ["", ""],
        ["비구독 고유 가맹점 수", df[df["category"] == "NonSub"]["merchant"].nunique()],
    ]

    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 20

    for i, (lbl, val) in enumerate(stats, 1):
        ca = ws3.cell(row=i, column=1, value=lbl)
        cb = ws3.cell(row=i, column=2, value=val)
        if i == 1:
            style_header(ca, fill=header_fill_b)
            style_header(cb, fill=header_fill_b)
        elif lbl in ("금액 통계", "케이스 분포 (거래 건수 기준)"):
            style_header(ca, fill=header_fill_l)
            style_header(cb, fill=header_fill_l)
        elif lbl == "":
            pass
        else:
            style_cell(ca, align=left)
            style_cell(cb, align=right)
            if isinstance(val, int):
                cb.number_format = '#,##0'

    wb.save(output_path)
    print(f"    엑셀 저장 완료: {output_path}")


# ──────────────────────────────────────────
# 섹션 7. SSAFY 카드 결제 API 형식
# ──────────────────────────────────────────

class SSAFYHeader(BaseModel):
    apiName:                        str
    transmissionDate:               str   # YYYYMMDD
    transmissionTime:               str   # HHMMSS
    institutionCode:                str = "00100"
    fintechAppNo:                   str = "001"
    apiServiceCode:                 str
    institutionTransactionUniqueNo: str   # 20자리
    apiKey:                         str
    userKey:                        str


class CardTransactionRequest(BaseModel):
    """createCreditCardTransaction API 요청 형식"""
    Header:         SSAFYHeader
    cardNo:         str   # 16자리 카드번호
    cvc:            str   # 3자리 CVC
    merchantId:     str   # 가맹점 ID (Long → str 전송)
    paymentBalance: str   # 거래금액 (Long → str 전송)


_seq_counter = 0

def _unique_no() -> str:
    global _seq_counter
    _seq_counter += 1
    return datetime.datetime.now().strftime("%Y%m%d%H%M%S") + str(_seq_counter).zfill(6)


def to_card_request(
    record:      dict,
    card_no:     str,
    cvc:         str,
    merchant_id: str,
    api_key:     str,
    user_key:    str,
) -> CardTransactionRequest:
    txn_date = record["date"]
    if isinstance(txn_date, datetime.date):
        date_str = txn_date.strftime("%Y%m%d")
        time_str = (f"{random.randint(0,23):02d}"
                    f"{random.randint(0,59):02d}"
                    f"{random.randint(0,59):02d}")
    else:
        now      = datetime.datetime.now()
        date_str = now.strftime("%Y%m%d")
        time_str = now.strftime("%H%M%S")

    header = SSAFYHeader(
        apiName="createCreditCardTransaction",
        transmissionDate=date_str,
        transmissionTime=time_str,
        apiServiceCode="createCreditCardTransaction",
        institutionTransactionUniqueNo=_unique_no(),
        apiKey=api_key,
        userKey=user_key,
    )
    return CardTransactionRequest(
        Header=header,
        cardNo=card_no,
        cvc=cvc,
        merchantId=merchant_id,
        paymentBalance=str(record["amount_krw"]),
    )


# ──────────────────────────────────────────
# 섹션 8. API 호출
# ──────────────────────────────────────────

API_URL = os.getenv(
    "FINANCE_CARD_API_URL",
    "https://finance-api.example.com/v1/credit-card/transactions",
)

def call_card_transaction(req: CardTransactionRequest) -> dict:
    resp = requests.post(API_URL, json=req.model_dump(), timeout=10)
    resp.raise_for_status()
    return resp.json()


# ──────────────────────────────────────────
# 섹션 9. 메인 파이프라인
# ──────────────────────────────────────────

def run_pipeline(
    n_users:       int  = 60,
    generate_only: bool = True,
    excel_path:    str  = "synth_transactions.xlsx",
    card_no:       str  = "",
    cvc:           str  = "",
    merchant_id:   str  = "",
    api_key:       str  = "",
    user_key:      str  = "",
):
    """
    generate_only=True  → 카드 정보 없이 데이터 생성 + 엑셀 저장
    generate_only=False → 엑셀 저장 후 실제 카드 결제 API 호출
    """
    print(f"[1] 유저 프로필 생성 (n_users={n_users}) ...")
    user_profiles = build_all_user_profiles(n_users)
    case_dist_users = Counter(p["case"] for p in user_profiles.values())
    print(f"    유저 케이스 분포: { {k: case_dist_users[k] for k in sorted(case_dist_users)} }")

    print("[2] 거래 구조 생성 ...")
    base_df = generate_base_transactions(n_users)
    print(f"    총 {len(base_df)}건 생성")

    print("[3] 한국 결제명 + 금액 후처리 ...")
    records = postprocess(base_df, user_profiles)

    sub_cnt = sum(1 for r in records if r["category"] == "Subscription")
    print(f"    구독: {sub_cnt}건 / 비구독: {len(records)-sub_cnt}건")

    print(f"[4] 엑셀 저장 ({excel_path}) ...")
    export_to_excel(records, excel_path)

    if generate_only:
        print("\n[생성 완료] generate_only=True 모드")
        print("  → 엑셀에서 데이터 확인 후 아래 항목 채워서 재실행하세요:")
        print("      card_no, cvc, merchant_id, api_key, user_key")
        print("      generate_only=False")
        return records

    # ── API 호출 모드 ──
    if not all([card_no, cvc, merchant_id, api_key, user_key]):
        raise ValueError("generate_only=False 시 card_no, cvc, merchant_id, api_key, user_key 모두 필요")

    print(f"[5] 카드 결제 API 호출 ({len(records)}건) ...")
    ok, fail = 0, 0
    for record in records:
        req = to_card_request(record, card_no, cvc, merchant_id, api_key, user_key)
        try:
            call_card_transaction(req)
            ok += 1
        except Exception as e:
            fail += 1
            if fail <= 5:
                print(f"    [FAIL] {record['merchant']}: {e}")

    print(f"[완료] 성공 {ok}건 / 실패 {fail}건")
    return records


# ──────────────────────────────────────────
# 실행
# ──────────────────────────────────────────

if __name__ == "__main__":

    # ── STEP 1: 카드 정보 없이 데이터 먼저 생성 ──
    # n_users=60: 케이스 6종(A~F) × 10명 균등 배분
    run_pipeline(
        n_users=60,
        generate_only=True,
        excel_path="synth_transactions.xlsx",
    )

    # ── STEP 2: 백엔드에서 카드 정보 받은 후 아래 주석 해제하고 재실행 ──
    # run_pipeline(
    #     n_users=60,
    #     generate_only=False,
    #     excel_path="synth_transactions.xlsx",
    #     card_no="1005518816096479",   # 백엔드에서 받은 카드번호
    #     cvc="725",                   # 백엔드에서 받은 CVC
    #     merchant_id="1",             # 백엔드에서 받은 가맹점 ID
    #     api_key="YOUR_API_KEY",
    #     user_key="YOUR_USER_KEY",
    # )
